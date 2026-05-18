"""
车辆配载与3D可视化装箱系统
==========================
基于 Streamlit + Plotly + Matplotlib + Openpyxl
3D Bin Packing (Extreme Point 启发式) + 物流行业约束

运行方式:
    cd ongoing && streamlit run app.py

依赖安装:
    pip install streamlit pandas numpy plotly matplotlib openpyxl
"""

# ==================== 一、导入依赖 ====================
import streamlit as st
import pandas as pd
import numpy as np
import plotly.graph_objects as go
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from dataclasses import dataclass, field
from typing import List, Tuple, Optional, Dict
import math

# ==================== 二、全局配置参数 ====================
# 尺寸缓冲 (Buffer): 考虑货物包装 / 托盘 / 防撞材料
BUFFER_FIXED_BOTTOM = 0       # Fixed_Bottom == '是' 时的缓冲 (cm)
BUFFER_UNFIXED = 0            # Fixed_Bottom == '否' 时的缓冲 (cm)

# 面积支撑比上限: 上层货物底面积不得超过下层支撑货物底面积的 N 倍
AREA_SUPPORT_RATIO = 4.0 / 3.0

# 平板车/高栏车高度上限 (cm): 无顶车型实际装载高度限制
OPEN_TOP_MAX_HEIGHT = 300.0

# 颜色方案 (用于 3D 和 2D 可视化)
COLOR_PALETTE = [
    '#1f77b4', '#ff7f0e', '#2ca02c', '#d62728', '#9467bd',
    '#8c564b', '#e377c2', '#7f7f7f', '#bcbd22', '#17becf',
    '#e41a1c', '#377eb8', '#4daf4a', '#984ea3', '#ff7f00',
    '#ffff33', '#a65628', '#f781bf', '#66c2a5', '#fc8d62',
]


# ==================== 三、数据结构定义 ====================

@dataclass
class CargoItem:
    """待装载的单个货物（已按 Quantity 展开 & Buffer 膨胀后）"""
    cargo_id: str           # 货物编号
    original_l: float       # 原始长 (cm)
    original_w: float       # 原始宽 (cm)
    original_h: float       # 原始高 (cm)
    dilated_l: float        # 膨胀后长 (cm) = 原始 + buffer
    dilated_w: float        # 膨胀后宽 (cm)
    dilated_h: float        # 膨胀后高 (cm)
    weight: float           # 单件毛重 (kg)
    fixed_bottom: bool      # True=固定底面, False=不固定
    specify_vehicle: str    # 指定车型: 箱车 / 飞翼车 / 平板车 / 高栏车 / 不指定
    unload_order: int       # 卸货顺序, 越小越先卸 (LIFO 用)
    max_stacking: int       # 最大可叠层数 (含本层, 1=不可叠放)
    quantity_idx: int = 0   # 同一货物 ID 的第几件 (0-based)

    @property
    def base_area(self) -> float:
        """底面积 = 膨胀长 × 膨胀宽 (cm²)"""
        return self.dilated_l * self.dilated_w

    @property
    def volume(self) -> float:
        """膨胀后体积 (cm³)"""
        return self.dilated_l * self.dilated_w * self.dilated_h

    @property
    def display_id(self) -> str:
        """显示用 ID, 多件货物追加 _序号"""
        if self.quantity_idx > 0:
            return f"{self.cargo_id}_{self.quantity_idx}"
        return self.cargo_id


@dataclass
class PlacedItem:
    """已装载进车厢的货物快照"""
    cargo_item: CargoItem
    x: float
    y: float
    z: float
    placed_l: float         # 放置后 X 方向实际占用长度
    placed_w: float         # Y 方向占用宽度
    placed_h: float         # Z 方向占用高度

    @property
    def x2(self): return self.x + self.placed_l
    @property
    def y2(self): return self.y + self.placed_w
    @property
    def z2(self): return self.z + self.placed_h


@dataclass
class VehicleTemplate:
    """车型模板（从 Excel 车型数据源读取）"""
    category: str           # 箱车 / 飞翼车 / 平板车 / 高栏车
    internal_l: float       # 内径长 (cm)
    internal_w: float       # 内径宽 (cm)
    internal_h: float       # 内径高 (cm), 平板/高栏受 OPEN_TOP_MAX_HEIGHT 限制
    max_payload: float      # 最大载重量 (kg)

    @property
    def is_open_top(self) -> bool:
        """平板车 / 高栏车（无物理顶盖, 但受 OPEN_TOP_MAX_HEIGHT 限制）"""
        return self.category in ('平板车', '高栏车')

    @property
    def display_name(self) -> str:
        """格式化: '9.6m 箱车'"""
        lm = self.internal_l / 100.0
        return f"{lm:.1f}m {self.category}"


@dataclass
class VehicleInstance:
    """正在执行装箱的车厢实例"""
    template: VehicleTemplate
    placed_items: List[PlacedItem] = field(default_factory=list)

    @property
    def display_name(self) -> str:
        return self.template.display_name

    @property
    def is_open_top(self) -> bool:
        return self.template.is_open_top

    @property
    def used_payload(self) -> float:
        return sum(p.cargo_item.weight for p in self.placed_items)

    @property
    def remaining_payload(self) -> float:
        return self.template.max_payload - self.used_payload

    @property
    def utilization(self) -> float:
        if self.template.max_payload <= 0:
            return 0.0
        return self.used_payload / self.template.max_payload


# ==================== 四、Excel 模板生成 ====================

def _apply_header_style(ws, headers, fill_color):
    """统一样式: 蓝色表头 + 边框"""
    header_fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True, size=11)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'),
    )
    for col, text in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=text)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border


def create_cargo_template_bytes() -> BytesIO:
    """生成「货物数据源」Excel 模板 (含示例数据 + 填写说明)"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '货物数据源'

    headers = [
        'Cargo_ID', 'Length(cm)', 'Width(cm)', 'Height(cm)',
        'Gross_Weight(kg)', 'Fixed_Bottom', 'Specify_Vehicle_Type',
        'Unload_Order', 'Max_Stacking_Layers', 'Quantity',
    ]
    _apply_header_style(ws, headers, '4472C4')

    sample = [
        ['A001', 120, 80, 100, 500, '否', '箱车', 1, 3, 5],
        ['A002', 200, 150, 50, 800, '是', '箱车', 2, 2, 2],
        ['B001', 100, 100, 80, 300, '否', '平板车', 1, 4, 10],
        ['B002', 180, 120, 90, 600, '否', '不指定', 3, 3, 3],
        ['C001', 150, 100, 120, 400, '是', '飞翼车', 1, 2, 4],
    ]
    thin = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'),
    )
    for ri, row in enumerate(sample, 2):
        for ci, v in enumerate(row, 1):
            cell = ws.cell(row=ri, column=ci, value=v)
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin

    widths = [12, 14, 14, 14, 18, 16, 22, 14, 22, 10]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    # 填写说明 sheet
    ws2 = wb.create_sheet('填写说明')
    instr = [
        ['字段', '说明', '可选值'],
        ['Cargo_ID', '货物唯一编号', '任意文本'],
        ['Length/Width/Height(cm)', '货物长宽高, 单位厘米', '正数'],
        ['Gross_Weight(kg)', '单件毛重, 单位公斤', '正数'],
        ['Fixed_Bottom', '是否固定底面(不可倒置/侧放)', '是 / 否'],
        ['Specify_Vehicle_Type', '指定所需车型', '箱车 / 飞翼车 / 平板车 / 高栏车 / 不指定'],
        ['Unload_Order', '卸货顺序, 越小越先卸(LIFO)', '正整数'],
        ['Max_Stacking_Layers', '最大可叠层数', '正整数'],
        ['Quantity', '件数', '正整数'],
    ]
    for ri, row in enumerate(instr, 1):
        for ci, v in enumerate(row, 1):
            c = ws2.cell(row=ri, column=ci, value=v)
            if ri == 1:
                c.font = Font(bold=True)
    ws2.column_dimensions['A'].width = 28
    ws2.column_dimensions['B'].width = 45
    ws2.column_dimensions['C'].width = 45

    out = BytesIO()
    wb.save(out)
    out.seek(0)
    return out


def create_vehicle_template_bytes() -> BytesIO:
    """生成「车型数据源」Excel 模板"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '车型数据源'

    headers = [
        'Vehicle_Category', 'Internal_Length(cm)',
        'Internal_Width(cm)', 'Internal_Height(cm)', 'Max_Payload(kg)',
    ]
    _apply_header_style(ws, headers, '548235')

    sample = [
        ['箱车', 960, 245, 270, 18000],
        ['箱车', 680, 240, 260, 10000],
        ['飞翼车', 960, 245, 260, 16000],
        ['平板车', 1370, 250, 0, 30000],
        ['高栏车', 960, 245, 0, 18000],
    ]
    thin = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'),
    )
    for ri, row in enumerate(sample, 2):
        for ci, v in enumerate(row, 1):
            c = ws.cell(row=ri, column=ci, value=v)
            c.alignment = Alignment(horizontal='center')
            c.border = thin

    for i, w in enumerate([22, 22, 22, 22, 18], 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    ws2 = wb.create_sheet('填写说明')
    instr = [
        ['字段', '说明', '可选值'],
        ['Vehicle_Category', '车型分类', '箱车 / 飞翼车 / 平板车 / 高栏车'],
        ['Internal_Length(cm)', '内径长度', '数值'],
        ['Internal_Width(cm)', '内径宽度', '数值'],
        ['Internal_Height(cm)', '内径高度, 平板/高栏填0', '数值'],
        ['Max_Payload(kg)', '最大载重量', '数值'],
    ]
    for ri, row in enumerate(instr, 1):
        for ci, v in enumerate(row, 1):
            c = ws2.cell(row=ri, column=ci, value=v)
            if ri == 1:
                c.font = Font(bold=True)
    ws2.column_dimensions['A'].width = 28
    ws2.column_dimensions['B'].width = 45
    ws2.column_dimensions['C'].width = 45

    out = BytesIO()
    wb.save(out)
    out.seek(0)
    return out


# ==================== 五、数据解析 ====================

def _map_columns(df: pd.DataFrame, keywords: Dict[str, List[str]]) -> Dict[str, str]:
    """根据关键词匹配 DataFrame 列名, 兼容中英文"""
    mapping: Dict[str, str] = {}
    for col in df.columns:
        cl = str(col).strip().lower().replace(' ', '_').replace('(', '').replace(')', '')
        for key, words in keywords.items():
            if key in mapping:
                continue
            for w in words:
                if w in cl:
                    mapping[key] = col
                    break
    return mapping


def parse_cargo_data(df: pd.DataFrame) -> List[CargoItem]:
    """
    解析货物 DataFrame, 按 Quantity 展开, 并应用 Buffer 膨胀.
    """
    kw = {
        'cargo_id':       ['cargo_id', '货物编号', '货号'],
        'length':         ['length', '长'],
        'width':          ['width', '宽'],
        'height':         ['height', '高'],
        'weight':         ['weight', '毛重', '重量', 'gross'],
        'fixed_bottom':   ['fixed_bottom', '固定底面', 'fixed'],
        'specify_vehicle':['specify', '指定车型', 'vehicle_type'],
        'unload_order':   ['unload', '卸货', '顺序'],
        'max_stacking':   ['stacking', '叠'],
        'quantity':       ['quantity', '数量'],
    }
    col = _map_columns(df, kw)

    items: List[CargoItem] = []
    for _, row in df.iterrows():
        try:
            cid = str(row.get(col.get('cargo_id', ''), ''))
            if not cid or cid == 'nan':
                continue
            l = float(row.get(col.get('length'), 0))
            w = float(row.get(col.get('width'), 0))
            h = float(row.get(col.get('height'), 0))
            wt = float(row.get(col.get('weight'), 0))
            fb_str = str(row.get(col.get('fixed_bottom'), '否')).strip()
            fb = fb_str in ('是', 'yes', 'YES', 'True', 'true', '1')
            sv = str(row.get(col.get('specify_vehicle'), '不指定')).strip()
            uo = int(float(row.get(col.get('unload_order'), 99)))
            ms = int(float(row.get(col.get('max_stacking'), 3)))
            qty = int(float(row.get(col.get('quantity'), 1)))

            buf = BUFFER_FIXED_BOTTOM if fb else BUFFER_UNFIXED

            for i in range(qty):
                items.append(CargoItem(
                    cargo_id=cid,
                    original_l=l, original_w=w, original_h=h,
                    dilated_l=l + buf, dilated_w=w + buf, dilated_h=h + buf,
                    weight=wt,
                    fixed_bottom=fb,
                    specify_vehicle=sv,
                    unload_order=uo,
                    max_stacking=ms,
                    quantity_idx=i + 1 if qty > 1 else 0,
                ))
        except Exception:
            continue
    return items


def parse_vehicle_data(df: pd.DataFrame) -> List[VehicleTemplate]:
    """解析车型数据, 平板车/高栏车高度上限为 OPEN_TOP_MAX_HEIGHT."""
    kw = {
        'category': ['category', '车型'],
        'length':   ['length', '长'],
        'width':    ['width', '宽'],
        'height':   ['height', '高'],
        'payload':  ['payload', '载重', 'max_'],
    }
    col = _map_columns(df, kw)

    templates: List[VehicleTemplate] = []
    for _, row in df.iterrows():
        try:
            cat = str(row.get(col.get('category'), '')).strip()
            if not cat or cat == 'nan':
                continue
            il = float(row.get(col.get('length'), 0))
            iw = float(row.get(col.get('width'), 0))
            ih = float(row.get(col.get('height'), 0))
            mp = float(row.get(col.get('payload'), 0))

            # 平板车 / 高栏车: 高度受 OPEN_TOP_MAX_HEIGHT 限制
            if cat in ('平板车', '高栏车'):
                ih = OPEN_TOP_MAX_HEIGHT
            templates.append(VehicleTemplate(
                category=cat, internal_l=il, internal_w=iw,
                internal_h=ih, max_payload=mp,
            ))
        except Exception:
            continue
    return templates


# ==================== 六、核心装箱算法 (3D-BPP Extreme Point) ====================

class BinPacker:
    """
    3D Bin Packing 装箱器
    使用 Extreme Point 启发式 + 贪婪放置策略.
    """

    def __init__(self):
        pass

    # ---- 6.1 旋转方向枚举 ----

    def get_orientations(self, item: CargoItem) -> List[Tuple[float, float, float]]:
        """
        返回允许的放置方向列表, 每个元素为 (l, w, h):
        - Fixed_Bottom='是': 仅水平旋转 (长宽互换), Z 轴固定为 dilated_h
        - Fixed_Bottom='否': 6 种全量旋转
        """
        dl, dw, dh = item.dilated_l, item.dilated_w, item.dilated_h

        if item.fixed_bottom:
            # 按宽度降序排列: 优先让宽面贴Y轴, 有助于填满车厢宽度
            candidates = sorted([(dl, dw, dh), (dw, dl, dh)], key=lambda o: o[1], reverse=True)
        else:
            candidates = [
                (dl, dw, dh), (dw, dl, dh),   # Z = dh
                (dl, dh, dw), (dh, dl, dw),   # Z = dw
                (dw, dh, dl), (dh, dw, dl),   # Z = dl
            ]

        # 去重（正方体货物可能有重复方向）
        seen = set()
        unique: List[Tuple[float, float, float]] = []
        for o in candidates:
            if o not in seen:
                seen.add(o)
                unique.append(o)
        return unique

    # ---- 6.2 几何碰撞 / 边界 / 堆叠检查 ----

    @staticmethod
    def _boxes_overlap(x1, y1, z1, l1, w1, h1,
                       x2, y2, z2, l2, w2, h2) -> bool:
        """判断两个轴对齐长方体是否相交"""
        return not (
            x1 + l1 <= x2 or x2 + l2 <= x1 or
            y1 + w1 <= y2 or y2 + w2 <= y1 or
            z1 + h1 <= z2 or z2 + h2 <= z1
        )

    @staticmethod
    def _check_bounds(x, y, z, l, w, h, veh: VehicleInstance) -> bool:
        """检查是否超出车厢内径边界 (所有车型统一校验 X/Y/Z)"""
        if x + l > veh.template.internal_l + 0.1:
            return False
        if y + w > veh.template.internal_w + 0.1:
            return False
        if z + h > veh.template.internal_h + 0.1:
            return False
        return True

    @staticmethod
    def _check_overlap(x, y, z, l, w, h, veh: VehicleInstance) -> bool:
        """检查与已放置货物是否重叠"""
        for p in veh.placed_items:
            if BinPacker._boxes_overlap(
                x, y, z, l, w, h,
                p.x, p.y, p.z, p.placed_l, p.placed_w, p.placed_h,
            ):
                return False
        return True

    @staticmethod
    def _find_supporters(x, y, z, l, w, h,
                         veh: VehicleInstance) -> List[PlacedItem]:
        """
        找出所有"正下方支撑货物":
        条件: 1) 下层货物顶面 == 当前货物底面 z;
              2) X-Y 投影有重叠.
        """
        if z < 0.01:            # 放在车厢地板上
            return []

        supporters: List[PlacedItem] = []
        eps = 0.1
        for p in veh.placed_items:
            if abs(p.z2 - z) > eps:
                continue
            x_over = min(x + l, p.x2) - max(x, p.x)
            y_over = min(y + w, p.y2) - max(y, p.y)
            if x_over > 0.01 and y_over > 0.01:
                supporters.append(p)
        return supporters

    @staticmethod
    def _get_stack_depth(x, y, z, l, w, h, veh: VehicleInstance) -> int:
        """
        递归计算当前放置位置下方已堆叠的层数。
        底板为第1层, 底板上的第一件货物为第2层, 依此类推。
        """
        if z < 0.01:
            return 1
        supporters = BinPacker._find_supporters(x, y, z, l, w, h, veh)
        if not supporters:
            return 1
        max_depth = 0
        for sup in supporters:
            d = BinPacker._get_stack_depth(
                sup.x, sup.y, sup.z,
                sup.placed_l, sup.placed_w, sup.placed_h,
                veh
            )
            if d > max_depth:
                max_depth = d
        return max_depth + 1

    @staticmethod
    def _check_stacking(item: CargoItem, supporters: List[PlacedItem]) -> bool:
        """
        堆叠物理规则:
        - 重不压轻: 上层重量 <= 下层重量
        - 面积支撑: 上层底面积 <= 下层底面积 × AREA_SUPPORT_RATIO
        """
        if not supporters:
            return True          # 地板支撑, 永远通过

        upper_area = item.base_area
        for sup in supporters:
            sup_item = sup.cargo_item
            if item.weight > sup_item.weight + 0.01:
                return False
            if upper_area > sup_item.base_area * AREA_SUPPORT_RATIO + 0.01:
                return False
        return True

    # ---- 6.3 候选放置点 (Extreme Points) ----

    @staticmethod
    def _get_candidate_points(veh: VehicleInstance) -> List[Tuple[float, float, float]]:
        """
        从已放置货物角点生成候选放置位置.
        每个已放置货物贡献 7 个角点 (3 个轴投影 + 3 个对角 + 1 个远端角).
        """
        pts = {(0.0, 0.0, 0.0)}           # 车厢原点

        for p in veh.placed_items:
            x, y, z = p.x, p.y, p.z
            l, w, h = p.placed_l, p.placed_w, p.placed_h
            for pt in [
                (x + l, y, z),             # 右侧, 同高
                (x, y + w, z),             # 前侧, 同高
                (x, y, z + h),             # 正上方
                (x + l, y + w, z),         # 右前角, 同高
                (x + l, y, z + h),         # 右侧上方
                (x, y + w, z + h),         # 前侧上方
                (x + l, y + w, z + h),     # 远端顶角
            ]:
                pts.add(pt)

        # 过滤掉在已放置货物内部的点, 以及超出车厢的点
        valid: List[Tuple[float, float, float]] = []
        tl, tw = veh.template.internal_l, veh.template.internal_w
        th = veh.template.internal_h

        for px, py, pz in pts:
            # 超出车厢
            if px > tl + 1 or py > tw + 1:
                continue
            if not veh.is_open_top and pz > th + 1:
                continue
            # 是否在某个已放货物内部
            inside = False
            for p in veh.placed_items:
                if (p.x < px < p.x2 - 0.01 and
                    p.y < py < p.y2 - 0.01 and
                    p.z < pz < p.z2 - 0.01):
                    inside = True
                    break
            if not inside:
                valid.append((px, py, pz))

        # 排序: 优先 Z 最低 → X+Y 最小 (从低往高、从里往外填充; 有利于 LIFO)
        valid.sort(key=lambda pt: (pt[2], pt[0] + pt[1]))
        return valid

    # ---- 6.4 单件货物放置 ----

    def try_place(self, item: CargoItem, veh: VehicleInstance) -> Optional[PlacedItem]:
        """
        尝试将货物放入车厢. 成功返回 PlacedItem, 失败返回 None.
        """
        # 重量预检
        if item.weight > veh.remaining_payload + 0.01:
            return None

        orientations = self.get_orientations(item)
        candidates = self._get_candidate_points(veh)

        best = None
        best_score = float('inf')

        for px, py, pz in candidates:
            for l, w, h in orientations:
                if not self._check_bounds(px, py, pz, l, w, h, veh):
                    continue
                if not self._check_overlap(px, py, pz, l, w, h, veh):
                    continue
                supporters = self._find_supporters(px, py, pz, l, w, h, veh)
                if pz > 0.01 and not supporters:
                    continue                # 浮空, 不允许
                if not self._check_stacking(item, supporters):
                    continue
                # ---- max_stacking 层数约束 ----
                stack_depth = self._get_stack_depth(px, py, pz, l, w, h, veh)
                if stack_depth > item.max_stacking:
                    continue
                # 检查支撑货物是否还能承受上方堆叠
                sup_ok = True
                for sup in supporters:
                    if stack_depth > sup.cargo_item.max_stacking:
                        sup_ok = False
                        break
                if not sup_ok:
                    continue

                # 评分: Z 优先 → X 优先 → Y 优先 (越低越好)
                # 贴边奖励: 贴Y=0或Y=W 的放置减少碎片空间, 奖励50
                edge_bonus = 50 if (py < 0.01 or abs(py + w - veh.template.internal_w) < 0.01) else 0
                score = pz * 10000 + px * 100 + py - edge_bonus
                if score < best_score:
                    best_score = score
                    best = (px, py, pz, l, w, h)

        if best is None:
            return None

        x, y, z, l, w, h = best
        placed = PlacedItem(cargo_item=item, x=x, y=y, z=z,
                            placed_l=l, placed_w=w, placed_h=h)
        veh.placed_items.append(placed)
        return placed

    # ---- 6.5 整车 / 多车装箱 ----

    def _template_volume(self, t: VehicleTemplate) -> float:
        """计算模板容积 (cm³), 用于排序"""
        h = t.internal_h if not t.is_open_top else min(t.internal_h, OPEN_TOP_MAX_HEIGHT)
        return t.internal_l * t.internal_w * h

    def _sort_items(self, items: List[CargoItem], category: str):
        """就地排序: 箱车按 LIFO 降序, 其余按体积降序"""
        if category == '箱车':
            items.sort(key=lambda it: (it.unload_order, it.volume), reverse=True)
        else:
            items.sort(key=lambda it: it.volume, reverse=True)

    def _pack_multi(self, items: List[CargoItem],
                    templates: List[VehicleTemplate]) -> Tuple[
                        List[VehicleInstance], List[CargoItem]]:
        """
        多车贪婪装箱 (兜底策略, 大车优先):
        按车型分组, 每组货物优先填已有车辆, 创建新车时从大到小尝试.
        """
        if not items or not templates:
            return [], items[:]

        # 车型目录, 按容积降序 (大车优先, 减少车辆总数)
        tpl_by_cat: Dict[str, List[VehicleTemplate]] = {}
        for t in templates:
            tpl_by_cat.setdefault(t.category, []).append(t)
        for cat in tpl_by_cat:
            tpl_by_cat[cat].sort(key=self._template_volume, reverse=True)
        all_tpls = sorted(templates, key=self._template_volume, reverse=True)

        groups: Dict[str, List[CargoItem]] = {}
        for item in items:
            sv = item.specify_vehicle
            if sv not in tpl_by_cat and sv != '不指定':
                sv = '不指定'
            groups.setdefault(sv, []).append(item)

        vehicles: List[VehicleInstance] = []
        unplaced: List[CargoItem] = []
        ordered_groups = sorted(groups.keys(),
                                key=lambda g: (0 if g == '不指定' else 1),
                                reverse=True)

        for sv in ordered_groups:
            grp = groups[sv]
            self._sort_items(grp, sv)
            candidate_tpls = all_tpls if sv == '不指定' else tpl_by_cat.get(sv, all_tpls)

            for item in grp:
                placed = False
                for veh in vehicles:
                    if sv != '不指定' and veh.template.category != sv:
                        continue
                    if self.try_place(item, veh):
                        placed = True
                        break

                if not placed:
                    for cand_tpl in candidate_tpls:
                        max_dim = max(item.dilated_l, item.dilated_w,
                                      item.dilated_h)
                        if (cand_tpl.internal_l < max_dim * 0.5 or
                            cand_tpl.internal_w < max_dim * 0.5 or
                            cand_tpl.internal_h < max_dim * 0.5):
                            continue
                        if item.weight > cand_tpl.max_payload:
                            continue
                        new_veh = VehicleInstance(template=cand_tpl)
                        if self.try_place(item, new_veh):
                            vehicles.append(new_veh)
                            placed = True
                            break

                if not placed:
                    unplaced.append(item)
        return vehicles, unplaced

    def pack(self, items: List[CargoItem],
             templates: List[VehicleTemplate]) -> Tuple[
                 List[VehicleInstance], List[CargoItem]]:
        """
        主装箱流程:
        阶段1: 尝试用一辆车装下全部兼容货物 (从小车型到大车型).
               先试 6.8m, 装不下再试 9.6m, …以此类推.
        阶段2: 若所有单车都装不下 → 多车贪婪 (大车优先).
        """
        if not items or not templates:
            return [], items[:]

        tpl_by_cat: Dict[str, List[VehicleTemplate]] = {}
        for t in templates:
            tpl_by_cat.setdefault(t.category, []).append(t)

        all_tpls_asc = sorted(templates, key=self._template_volume)  # 小→大

        # ---- 阶段1: 尝试一辆车装下所有兼容货物 ----
        for tpl in all_tpls_asc:
            compatible: List[CargoItem] = []
            incompatible: List[CargoItem] = []
            for item in items:
                sv = item.specify_vehicle
                if sv in ('不指定', tpl.category) or sv not in tpl_by_cat:
                    compatible.append(item)
                else:
                    incompatible.append(item)

            if not compatible:
                continue

            self._sort_items(compatible, tpl.category)

            veh = VehicleInstance(template=tpl)
            failed = False
            for item in compatible:
                if not self.try_place(item, veh):
                    failed = True
                    break

            if not failed and len(veh.placed_items) == len(compatible):
                # 一辆车全装下!
                if not incompatible:
                    return [veh], []
                # 递归处理不兼容货物
                more_v, more_u = self.pack(incompatible, templates)
                return [veh] + more_v, more_u

        # ---- 阶段2: 单车装不下 → 多车兜底 ----
        return self._pack_multi(items, templates)


# ==================== 七、可视化模块 ====================

def _box_mesh_traces(x, y, z, l, w, h, color, name='', opacity=0.85):
    """生成一个长方体的 12 个三角面 (Mesh3d trace 列表)"""
    v = [
        (x, y, z), (x + l, y, z), (x + l, y + w, z), (x, y + w, z),      # 0-3 底面
        (x, y, z + h), (x + l, y, z + h),                                # 4-5
        (x + l, y + w, z + h), (x, y + w, z + h),                        # 6-7 顶面
    ]
    # 2 个三角面 per 面 → 12 个三角形
    faces = [
        (0, 1, 2), (0, 2, 3),     # 底
        (4, 5, 6), (4, 6, 7),     # 顶
        (0, 1, 5), (0, 5, 4),     # 前 (Y=0 侧)
        (2, 3, 7), (2, 7, 6),     # 后 (Y=W 侧)
        (1, 2, 6), (1, 6, 5),     # 右 (X=L 侧)
        (0, 3, 7), (0, 7, 4),     # 左 (X=0 侧)
    ]
    traces = []
    for tri in faces:
        traces.append(go.Mesh3d(
            x=[v[i][0] for i in tri],
            y=[v[i][1] for i in tri],
            z=[v[i][2] for i in tri],
            color=color,
            opacity=opacity,
            flatshading=True,
            showlegend=False,
            name=name,
            hoverinfo='none',
        ))
    return traces


def create_3d_figure(vehicles: List[VehicleInstance]) -> go.Figure:
    """Plotly 3D 交互装箱图"""
    fig = go.Figure()
    ci = 0

    for veh in vehicles:
        tl, tw = veh.template.internal_l, veh.template.internal_w
        th = veh.template.internal_h
        if veh.is_open_top:
            # 给无顶车厢一个可视高度: max(货物顶面) + 20
            th = max((p.z2 for p in veh.placed_items), default=200) + 20

        # ---- 车厢半透明线框 ----
        corners = np.array([
            [0, 0, 0], [tl, 0, 0], [tl, tw, 0], [0, tw, 0],
            [0, 0, th], [tl, 0, th], [tl, tw, th], [0, tw, th],
        ])
        edges = [(0, 1), (1, 2), (2, 3), (3, 0),
                 (4, 5), (5, 6), (6, 7), (7, 4),
                 (0, 4), (1, 5), (2, 6), (3, 7)]
        for s, e in edges:
            fig.add_trace(go.Scatter3d(
                x=[corners[s][0], corners[e][0]],
                y=[corners[s][1], corners[e][1]],
                z=[corners[s][2], corners[e][2]],
                mode='lines',
                line=dict(color='rgba(120,120,120,0.45)', width=1.5),
                showlegend=False,
                hoverinfo='none',
            ))

        # ---- 货物方块 + 标签 ----
        for placed in veh.placed_items:
            color = COLOR_PALETTE[ci % len(COLOR_PALETTE)]
            ci += 1

            x, y, z = placed.x, placed.y, placed.z
            l, w, h = placed.placed_l, placed.placed_w, placed.placed_h

            for trace in _box_mesh_traces(x, y, z, l, w, h, color):
                fig.add_trace(trace)

            # 标签在货物顶面中心上方
            fig.add_trace(go.Scatter3d(
                x=[x + l / 2],
                y=[y + w / 2],
                z=[z + h + 3],
                mode='text',
                text=placed.cargo_item.display_id,
                textfont=dict(size=9, color='#222'),
                showlegend=False,
                hoverinfo='none',
            ))

    fig.update_layout(
        scene=dict(
            xaxis_title='X: 长度 (cm)',
            yaxis_title='Y: 宽度 (cm)',
            zaxis_title='Z: 高度 (cm)',
            aspectmode='data',
            camera=dict(eye=dict(x=1.6, y=1.6, z=1.1)),
        ),
        margin=dict(l=0, r=0, t=40, b=0),
        height=650,
        title='3D 装箱可视化',
    )
    return fig


def create_2d_views(vehicle: VehicleInstance, idx: int) -> plt.Figure:
    """工程三视图: 俯视(Top) / 正视(Front) / 侧视(Side)"""
    tl, tw = vehicle.template.internal_l, vehicle.template.internal_w
    th = vehicle.template.internal_h
    if vehicle.is_open_top:
        th = max((p.z2 for p in vehicle.placed_items), default=200) + 10

    fig, axes = plt.subplots(1, 3, figsize=(18, 5.5))

    configs = [
        ('俯视图 (Top) — X‑Y', 0, tl, 0, tw, 'X (cm)', 'Y (cm)'),
        ('正视图 (Front) — X‑Z', 0, tl, 0, th, 'X (cm)', 'Z (cm)'),
        ('侧视图 (Side) — Y‑Z', 0, tw, 0, th, 'Y (cm)', 'Z (cm)'),
    ]

    for ax_i, (title, x_min, x_max, y_min, y_max, xlbl, ylbl) in enumerate(configs):
        ax = axes[ax_i]
        ax.set_title(f'{title}\n车辆 #{idx+1}: {vehicle.display_name}', fontsize=10)
        ax.set_aspect('equal')
        ax.set_xlim(x_min, x_max)
        ax.set_ylim(y_min, y_max)
        ax.set_xlabel(xlbl)
        ax.set_ylabel(ylbl)

        # 车厢边界
        border = mpatches.Rectangle(
            (x_min, y_min), x_max - x_min, y_max - y_min,
            fill=False, edgecolor='gray', linewidth=1.5, linestyle='--',
        )
        ax.add_patch(border)

        ci = 0
        for p in vehicle.placed_items:
            col = COLOR_PALETTE[ci % len(COLOR_PALETTE)]
            ci += 1

            if ax_i == 0:       # 俯视
                rx, ry, rw, rh = p.x, p.y, p.placed_l, p.placed_w
            elif ax_i == 1:     # 正视
                rx, ry, rw, rh = p.x, p.z, p.placed_l, p.placed_h
            else:               # 侧视
                rx, ry, rw, rh = p.y, p.z, p.placed_w, p.placed_h

            rect = mpatches.Rectangle(
                (rx, ry), rw, rh,
                facecolor=col, edgecolor='black',
                alpha=0.75, linewidth=0.5,
            )
            ax.add_patch(rect)
            ax.text(rx + rw / 2, ry + rh / 2,
                    p.cargo_item.display_id,
                    ha='center', va='center', fontsize=5.5,
                    color='white', weight='bold')

        ax.grid(True, alpha=0.25)

    plt.tight_layout()
    return fig


# ==================== 八、Streamlit 界面 ====================

def check_password():
    """简单密码验证 — 密码通过环境变量 APP_PASSWORD 设置，默认 '123456'"""
    import os, hmac

    correct = os.environ.get('APP_PASSWORD', '123456')

    if 'authenticated' not in st.session_state:
        st.session_state.authenticated = False

    if st.session_state.authenticated:
        return True

    st.markdown('## 🔐 登录')
    pwd = st.text_input('请输入访问密码', type='password', key='login_pwd')
    if st.button('登录', type='primary'):
        if hmac.compare_digest(pwd, correct):
            st.session_state.authenticated = True
            st.rerun()
        else:
            st.error('密码错误')
    return False


def main():
    st.set_page_config(
        page_title='车辆配载与3D装箱系统',
        page_icon='🚛',
        layout='wide',
    )

    if not check_password():
        st.stop()

    st.title('🚛 车辆配载与 3D 可视化装箱系统')

    # ===== 侧边栏: 控制面板 =====
    with st.sidebar:
        st.header('📋 控制面板')

        st.subheader('1. 下载 Excel 模板')
        c1, c2 = st.columns(2)
        with c1:
            st.download_button(
                '📦 货物模板',
                data=create_cargo_template_bytes(),
                file_name='cargo_template.xlsx',
                mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            )
        with c2:
            st.download_button(
                '🚚 车型模板',
                data=create_vehicle_template_bytes(),
                file_name='vehicle_template.xlsx',
                mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            )

        st.markdown('---')
        st.subheader('2. 上传数据')
        cargo_f = st.file_uploader('上传货物 Excel', type=['xlsx', 'xls'], key='c')
        veh_f = st.file_uploader('上传车型 Excel', type=['xlsx', 'xls'], key='v')

        st.markdown('---')
        st.subheader('3. 执行计算')
        run = st.button('▶️ 开始装箱计算', type='primary', use_container_width=True)

        st.markdown('---')
        st.caption(f'Buffer(固定底面):   {BUFFER_FIXED_BOTTOM} cm')
        st.caption(f'Buffer(非固定):     {BUFFER_UNFIXED} cm')
        st.caption(f'面积支撑比上限:     {AREA_SUPPORT_RATIO:.2f}')
        st.caption(f'平板/高栏高度上限:  {OPEN_TOP_MAX_HEIGHT} cm')

    # ===== 主区域 =====
    if not run:
        st.info('👈 左侧下载模板 → 填写数据 → 上传 → 点击「开始装箱计算」')
        st.markdown('''
        ### 使用步骤
        1. **下载模板** — 左侧"货物模板"和"车型模板"
        2. **填写数据** — Excel 中填入实际数据（模板自带示例）
        3. **上传文件** — 将两个 Excel 上传
        4. **执行计算** — 点击「开始装箱计算」
        5. **查看结果** — 浏览 3D 视图 / 2D 三视图 / 明细表
        ''')
        return

    if cargo_f is None or veh_f is None:
        st.error('❌ 请同时上传货物数据和车型数据')
        return

    # ---- 加载数据 ----
    with st.spinner('加载数据中…'):
        try:
            cargo_df = pd.read_excel(cargo_f)
            veh_df = pd.read_excel(veh_f)
        except Exception as e:
            st.error(f'❌ Excel 读取失败: {e}')
            return

        cargo_items = parse_cargo_data(cargo_df)
        veh_tpls = parse_vehicle_data(veh_df)

    if not cargo_items:
        st.error('❌ 未解析到有效货物数据')
        return
    if not veh_tpls:
        st.error('❌ 未解析到有效车型数据')
        return

    st.success(f'✅ 加载成功: {len(cargo_items)} 件货物, {len(veh_tpls)} 种车型')

    # ---- 执行装箱 ----
    with st.spinner('3D 装箱计算中…'):
        packer = BinPacker()
        vehicles, unplaced = packer.pack(cargo_items, veh_tpls)

    st.markdown('---')

    # ---- 汇总看板 ----
    total = len(cargo_items)
    placed_cnt = sum(len(v.placed_items) for v in vehicles)
    m1, m2, m3, m4 = st.columns(4)
    m1.metric('总货物件数', total)
    m2.metric('已装载件数', placed_cnt)
    m3.metric('使用车辆数', len(vehicles))
    m4.metric('装载率', f'{placed_cnt / total * 100:.1f}%' if total else '0%')

    if unplaced:
        st.warning(f'⚠️ {len(unplaced)} 件未装载: '
                   f'{", ".join(it.display_id for it in unplaced)}')

    if not vehicles:
        st.error('未能装载任何货物, 请检查数据')
        return

    # ---- 车辆汇总表 ----
    st.subheader('📊 车辆装载汇总')
    summary = []
    for i, v in enumerate(vehicles):
        summary.append({
            '车辆': f'#{i+1}',
            '车型': v.display_name,
            '件数': len(v.placed_items),
            '总重(kg)': round(v.used_payload, 1),
            '最大载重(kg)': v.template.max_payload,
            '载重利用率': f'{v.utilization * 100:.1f}%',
        })
    st.dataframe(pd.DataFrame(summary), use_container_width=True)

    # ---- 货物明细 ----
    st.subheader('📦 货物装载明细')
    details = []
    for vi, v in enumerate(vehicles):
        for p in v.placed_items:
            details.append({
                '车辆': f'#{vi+1} {v.display_name}',
                '货物编号': p.cargo_item.display_id,
                '原始尺寸(cm)': f'{p.cargo_item.original_l}×{p.cargo_item.original_w}×{p.cargo_item.original_h}',
                '放置尺寸(cm)': f'{p.placed_l:.1f}×{p.placed_w:.1f}×{p.placed_h:.1f}',
                '坐标(X,Y,Z)': f'({p.x:.0f}, {p.y:.0f}, {p.z:.0f})',
                '重量(kg)': p.cargo_item.weight,
                '固定底面': '是' if p.cargo_item.fixed_bottom else '否',
                '卸货顺序': p.cargo_item.unload_order,
            })
    if details:
        st.dataframe(pd.DataFrame(details), use_container_width=True)

    # ---- 可视化 ----
    st.markdown('---')
    st.header('🎨 可视化')
    t3, t2 = st.tabs(['3D 交互视图', '2D 工程三视图'])

    with t3:
        st.caption('🖱 拖拽旋转 | 滚轮缩放 | 右键平移')
        fig3d = create_3d_figure(vehicles)
        st.plotly_chart(fig3d, use_container_width=True)

    with t2:
        for vi, v in enumerate(vehicles):
            st.markdown(f'### 车辆 #{vi + 1} — {v.display_name}')
            fig2d = create_2d_views(v, vi)
            st.pyplot(fig2d)
            buf = BytesIO()
            fig2d.savefig(buf, format='png', dpi=150, bbox_inches='tight')
            buf.seek(0)
            st.download_button(
                f'📥 下载三视图 #{vi+1}',
                data=buf,
                file_name=f'vehicle_{vi+1}_views.png',
                mime='image/png',
                key=f'dl2d_{vi}',
            )
            plt.close(fig2d)


if __name__ == '__main__':
    main()